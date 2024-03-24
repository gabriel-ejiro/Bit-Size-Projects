#!/usr/bin/env python
# coding: utf-8

# In[32]:


# creating_spreadsheet
from openpyxl import Workbook

def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)
    
if __name__ == "__main__":
    create_workbook(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[33]:


# adding_data.py

from openpyxl import Workbook

def create_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 42
    sheet['A2'] = "Arsenal"
    sheet['A3'] = 3.14159265359
    sheet['B1'] = 20
    sheet['B2'] = "Chelsea"
    workbook.save(path)
    
if __name__ == "__main__":
    create_workbook(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[34]:


# adding_rows.py
# best to use a list of lists

from openpyxl import Workbook

def create_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    data = [["Name", "Age", "City", "Sex"],
            ["Peter", 20, "London", "M"],
            ["Jane", 21, "Paris", "F"],
            ["Joe", 22, "New York", "M"]]
    for row in data:
        sheet.append(row)
    workbook.save(path)
    
if __name__ == "__main__":
    create_workbook(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[35]:


# create sheet title

from openpyxl import Workbook

def create_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MySheet"
    sheet2 = workbook.create_sheet("MySheet2")
    workbook.save(path)
    

if __name__ == "__main__":
    create_workbook(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[37]:


# delete sheets

import openpyxl

def create_worksheets(path):
    workbook = openpyxl.Workbook()
    workbook.create_sheet()
    print(workbook.sheetnames)
    # Insert a worksheet
    workbook.create_sheet(index=1, title="MySheet")
    print(workbook.sheetnames)
    del workbook["MySheet"]
    print(workbook.sheetnames)
    workbook.save(path)
    
if __name__ == "__main__":
    create_worksheets(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[45]:


#inserting columns

from openpyxl import Workbook

def inserting_cols_rows(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Hello"
    sheet['A2'] = "From"
    sheet['A3'] = "OpenPyXL"
    # insert a column before A
    sheet.insert_cols(idx=1)
    # insert 2 rows starting on the second row
    sheet.insert_rows(idx=2, amount=2)
    workbook.save(path)
    
    
if __name__ == "__main__":
    inserting_cols_rows("inserting.xlsx")


# In[47]:


# deleting columns and rows

from openpyxl import Workbook

def deleting_cols_rows(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Hello"
    sheet['A2'] = "From"
    sheet['A3'] = "OpenPyXL"
    # delete a column
    sheet.delete_cols(idx=1)
    # delete 2 rows on the second row
    sheet.delete_rows(idx=2, amount=2)
    workbook.save(path)
    
    
if __name__ == "__main__":
    deleting_cols_rows("deleting.xlsx")


# In[65]:


# editing cell data 

from openpyxl import load_workbook

def edit(path, data):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    for cell in data:
        current_value = sheet[cell].value
        sheet[cell] = current_value
        print(f"changing {cell} from {current_value} to {data[cell]}")
    workbook.save(path)
    
if __name__ == "__main__":
    data = {"A1": "Hello", "A2": "From", "A3": "OpenPyXL"}
    edit(r"C:\Users\ejiro\Downloads\edit.xlsx", data)


# In[67]:


# creating merged cell

from openpyxl import Workbook
from openpyxl.styles import Alignment

def create_merged_cells(path, value):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A2:E3")
    top_left_cell = sheet['A2']
    top_left_cell.alignment = Alignment(horizontal='center', vertical="center")
    
    sheet["A2"] = value
    workbook.save(path)
    
    
if __name__ == "__main__":
    create_merged_cells(r"C:\Users\ejiro\Downloads\hello.xlsx", "Hello")


# In[69]:


# folding rows and columns

from openpyxl import Workbook

def folding(path, rows=None, cols=None, hidden=True):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    if rows:
        begin_row, end_row = rows
        sheet.row_dimensions.group(begin_row, end_row, hidden=hidden)
    if cols:
        begin_col, end_col = cols
        sheet.column_dimensions.group(begin_col, end_col, hidden=hidden)
    workbook.save(path)
    
if __name__ == "__main__":
    folding(r"C:\Users\ejiro\Downloads\hello.xlsx", rows=(1, 3), cols=("C", "F"))


# In[83]:


# freezing panes
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def freeze(path, row_to_freeze):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Freeze"
    freeze_cell = f"A{row_to_freeze + 1}"  # Convert row number to cell coordinate
    sheet.freeze_panes = freeze_cell
    headers = ["Name", "Age", "City", "Sex"]
    sheet.append(headers)
    data = [("Mike", 20, "London", "M"), ("Jane", 21, "Paris", "F")]
    for d in data:
        sheet.append(d)
    
    workbook.save(path)
    
if __name__ == "__main__":
    freeze(r"C:\Users\ejiro\Downloads\hello.xlsx", row_to_freeze=1)


# In[87]:


# font_sizes

import openpyxl
from openpyxl.styles import Font

def font_demo(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    cell = sheet["A1"]
    cell.font = Font(size=12)
    cell.value = "Hello"
    
    cell2 = sheet["A2"]
    cell2.font = Font(name='Courier New', size=24, color="FF0000")
    sheet["A2"] = "OpenPyXL"
    
    workbook.save(path)
    
if __name__ == "__main__":
    font_demo(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[93]:


# Alignments (rotate text, set text wrapping, and for indentation)

from openpyxl import Workbook
from openpyxl.styles import Alignment

def center_text(path, horizontal="center", vertical="center"):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hellossss"
    sheet["A1"].alignment = Alignment(horizontal=horizontal, vertical=vertical)
    
    sheet["A2"] = "OpenPyXL"
    sheet["A5"] = "Hello World"
    sheet["A5"].alignment = Alignment(wrap_text=True)
    sheet["A7"] = "Hello"
    sheet["A7"].alignment = Alignment(text_rotation=90)
    
    workbook.save(path)
if __name__ == "__main__":
    center_text(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[97]:


# creating a border

from openpyxl import Workbook
from openpyxl.styles import Border, Side

def border(path):
    pink = "FFC0CB"
    green = "008000"
    thin = Side(border_style="thin", color=pink)
    double = Side(border_style="double", color=green)
    
    workbook = Workbook()
    sheet = workbook.active
    
    sheet["A1"] = "Hello"
    sheet["A1"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet["A2"] = "OpenPyXL"
    sheet["A2"].border = Border(top=double, left=double, right=double, bottom=double)
    workbook.save(path)
    

if __name__ == "__main__":
    border(r"C:\Users\ejiro\Downloads\hello.xlsx")
    


# In[101]:


# background color

from openpyxl import Workbook
from openpyxl.styles import PatternFill

def background_colors(path):
    workbook = Workbook()
    sheet = workbook.active
    yellow = "00FF00"
    for rows in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=12):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
                
    workbook.save(path)
    
if __name__ == "__main__":
    background_colors(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[ ]:


# inserting images into cells using add_image

from openpyxl import Workbook
from openpyxl.drawing.image import Image

def insert_image(path, image_path):
    workbook = Workbook()
    sheet = workbook.active
    img = Image(image_path)
    sheet.add_image(img, "A1")
    workbook.save(path)
    
    
if __name__ == "__main__":
    insert_image(r"C:\Users\ejiro\Downloads\hello.xlsx", r"C:\Users\ejiro\Downloads\hello.png")


# In[105]:


# styling merged_cells

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, GradientFill, Alignment

def merge_style(path):
    workbook = Workbook()
    sheet = workbook.active
    cell_range = "A1:B3"
    sheet.merge_cells(cell_range)
    top_left_cell = sheet["A1"]
    
    light_blue = "ADD8E6"
    green = "00FF00"
    thin = Side(border_style="thin", color=light_blue)
    double = Side(border_style="double", color=green)
    
    top_left_cell.value = "Hello Xavi"
    for column in sheet[cell_range]:
        for cell in column:
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            
    top_left_cell.fill = GradientFill(stop=("0000FF", "FFFFFF"))
    top_left_cell.font = Font(color="00FF00")
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
    workbook.save(path)
    
    
if __name__ == "__main__":
    merge_style(r"C:\Users\ejiro\Downloads\hello.xlsx")
    


# In[107]:


# using a named style

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, NamedStyle


def named_style(path):
    workbook = Workbook()
    sheet = workbook.active
    
    red = "FF0000"
    font = Font(bold=True, size=20)
    thick = Side(border_style="thick", color=red)
    border = Border(top=thick, left=thick, right=thick, bottom=thick)
    named_style = NamedStyle(name="test_style", font=font, border=border)
    
    sheet["A1"] = "Hello Xavi"
    sheet["A1"].style = named_style
    
    sheet["B1"] = "Bye Xavi"
    sheet["B1"].style = named_style
    
    cell_range = "A1:B3"
    sheet.merge_cells(cell_range)
    top_left_cell = sheet["A1"]
    
    workbook.save(path)
    
if __name__ == "__main__":
    named_style(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[110]:


# using_colorscale

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

def apply_colorscale(path):
    workbook = load_workbook(path)
    sheet = workbook.active
    
    
    red = "FF0000"
    green = "00FF00"
    black = "000000"
    
    color_scale_rule = ColorScaleRule(start_type="num",
                                start_value=1,
                                start_color=red,
                                mid_type="num",
                                mid_value=3,
                                mid_color=green,
                                end_type="num",
                                end_value=5,
                                end_color=green)
    sheet.conditional_formatting.add("A1:B5", color_scale_rule)
    workbook.save(path)
    
if __name__ == "__main__":
    apply_colorscale(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[ ]:


# using_iconset.py

from openpyxl import Workbook
from opennpyl.formatting.rule import IconSet, FormatObject, Rule

def applying_iconset(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active

    first = FormatObject(type="num", val=0)
    mid = FormatObject(type="num", val=3)
    last = FormatObject(type="num", val=5)
    
    icon_set = IconSet(iconSet="3Arrows", cfvo=[first, mid, last],
                        showValue="None", percent=None, reverse="true")
    rule = Rule(type="iconSet", iconSet=icon_set)
    sheet.conditional_formatting.add("A1:B5", rule)
    workbook.save(path)
    
if __name__ == "__main__":
    applying_iconset(r"C:\Users\ejiro\Downloads\hello.xlsx", output_path=r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[ ]:


# using iconsetrule (simplified)

from openpyxl import Workbook
from openpyxl.formatting.rule import IconSetRule

def applying_iconsetrule(path, output_path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    
    icon_set_rule = IconSetRule(type="5Arrows", "num", [1, 2, 3, 4, 5])
    sheet.conditional_formatting.add("A1:B5", icon_set_rule)
    workbook.save(output_path)
    
    
if __name__ == "__main__":
    applying_iconsetrule(r"C:\Users\ejiro\Downloads\hello.xlsx", output_path=r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[116]:


# creating charts

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

workbook = Workbook()
sheet = workbook.active

# add data to spreadsheet
data_rows = [
    ["Book", "Kindle", "Paperback"],
    [1, 9.99, 14.99],
    [2, 5.99, 9.99],
    [3, 19.99, 25.99],
    [4, 4.99, 8.99],
    [5, 14.99, 19.99],
]

for row in data_rows:
    sheet.append(row)
    
# create the bar chart
bar_chart = BarChart()
bar_chart.title = "Price Comparison"
bar_chart.style = 13
bar_chart.y_axis.title = "Price"
bar_chart.x_axis.title = "Book Number"


data = Reference(worksheet=sheet,
                min_row=1,
                max_row=10,
                min_col=2,
                max_col=3)

bar_chart.add_data(data, titles_from_data=True)
sheet.add_chart(bar_chart, "E2")

workbook.save(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[120]:


# creating a chartsheet

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

wb = Workbook()
ws = wb.active
chart_sheet = wb.create_chartsheet()

rows = [
    ["Python", 50],
    ["C++", 30],
    ["Java", 10],
    ["C#", 10],
    ["JavaScript", 10],
    
]

for row in rows:
    ws.append(row)
    
chart = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=5)
data = Reference(ws, min_col=2, min_row=1, max_row=5)

chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = "Programming Languages"
chart_sheet.add_chart(chart)
wb.save(r"C:\Users\ejiro\Downloads\hello.xlsx")


# In[123]:


# converting csv to excel

import csv
import openpyxl

def csv_to_excel(csv_path, excel_path):
    csv_data = []
    with open(csv_path) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)
            
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in csv_data:
        sheet.append(row)
    workbook.save(excel_path)
    
if __name__ == "__main__":
    csv_to_excel("books.csx", "converted_books.xlsx")


# In[124]:


# convert excel spreadsheet to csv

import csv
import openpyxl

def excel_to_csv(excel_path, csv_path):
    workbook = openpyxl.load_workbook(filename=excel_path)
    sheet = workbook.active
    csv_data = []
    
    # read data from Excel
    for value in sheet.iter_rows(values_only=True):
        csv_data.append(value)
        
    # write to csv
    with open(csv_path, "w") as csv_file_obj:
        writer = csv.writer(csv_file_obj, delimter=',')
        for line in csv_data:
            writer.writerow(row)
            
if __name__ == "__main__":
    excel_to_csv("books.xlsx", "converted_books.csv")


# In[126]:


# using pandas with excel

import pandas as pd

df = pd.read_excel(r"C:\Users\ejiro\Downloads\hello.xlsx")
print(df)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




